# Read excel files and compile income and expense for the mont
import openpyxl
import sys

income = sys.argv[1]
expense = sys.argv[2]
totExpense = 0
totSalary = 0

# Pass path as arg
path = "C:/Users/arvid/OneDrive/Dokument/Penningpung/transaktioner/transJan22_ut.xlsx"

# Open workbook and get active sheet from active attribute
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# test
print(income, expense)
print("test :")

# Phrases in payment description
FriendPay  = "Swish från"
Salary = "Insättning Lön/Pension"

# Iterate through the sheet from row 5 to last
for row in range(5, sheet_obj.max_row+1):
    description = sheet_obj.cell(row, column=2)
    amount = sheet_obj.cell(row, column=3)

    # Check for +- 0 expenses, paying for friend etc
    if FriendPay in description.value:
        totExpense += amount.value

    # Check for expenses, negative amounts
    elif amount.value < 0:
        totExpense += amount.value

    # Check for salary
    elif Salary in description.value:
        totSalary += amount.value


print("utgifter", totExpense)
print("lon", totSalary)
print("inkomst", totSalary+int(income))

wb_obj = openpyxl.load_workbook("C:/Users/arvid/OneDrive/Dokument/Penningpung/in_ut.xlsx")
sheet_obj = wb_obj.active

column = 2
cellIn = sheet_obj.cell(2, column)

while True:
    if cellIn.value is None:
        break
    print(cellIn.value)
    column += 1
    cellIn = sheet_obj.cell(2, column)



print(column)

cellEx = sheet_obj.cell(3, column)
cellTot = sheet_obj.cell(5, column)

cellIn.value = totSalary+int(income)
cellEx.value = totExpense
cellTot.value = totSalary+int(income)+totExpense

wb_obj.save("C:/Users/arvid/OneDrive/Dokument/Penningpung/in_ut.xlsx")